#!/usr/bin/env python3
"""
Generate an Excel financial model for DriveLedger (Drivly / DriveLedg) scenarios.
Creates `financial_model.xlsx` in the repository root with sheets:
- Assumptions
- Scenarios
- UnitEconomics

Requires: openpyxl
Run: python scripts/generate_financial_model.py
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = Workbook()

# --- Assumptions sheet ---
ass = wb.active
ass.title = 'Assumptions'
ass.append(['Key', 'Value', 'Notes'])
ass['A1'].font = Font(bold=True)
ass['B1'].font = Font(bold=True)
ass['C1'].font = Font(bold=True)

ass_data = [
    ('Country', 'Canada', 'Target market / country'),
    ('Currency', 'CAD', 'Currency for all values in this sheet'),
    ('TAM (total drivers)', 400000, 'Estimated total addressable market (Canada - gig/active drivers)'),
    ('Default SOM %', 0.05, 'Serviceable obtainable market %'),
    ('price_standard (CAD/yr)', 79, 'Standard annual price (CAD)'),
    ('price_pro (CAD/yr)', 149, 'Pro annual price (CAD)'),
    ('default_pro_pct', 0.70, 'Default percent of paid users on Pro'),
    ('CAC (CAD)', 70, 'Cost to acquire a paid user (CAD)'),
    ('Gross Margin', 0.80, 'Gross margin on subscription revenue'),
    ('Lifetime (years)', 2, 'Average paid user lifetime (years)'),
    ('growth_conservative', 0.10, 'Annual growth rate for paying users (conservative)'),
    ('growth_base', 0.30, 'Annual growth rate for paying users (base)'),
    ('growth_optimistic', 0.60, 'Annual growth rate for paying users (optimistic)'),
    ('annual_cost_growth', 0.03, 'Annual growth rate for operating costs'),
    ('DEV_ONE_OFF (CAD)', 450000, 'One-time development cost (Year 1, CAD)'),
    ('OPS_ANNUAL (CAD)', 300000, 'Annual operating cost (CAD)'),
]

r = 2
for key, val, note in ass_data:
    ass[f'A{r}'] = key
    ass[f'B{r}'] = val
    ass[f'C{r}'] = note
    r += 1

# Adjust column widths
for col in ['A','B','C']:
    ass.column_dimensions[col].width = 28

# --- Scenarios sheet ---
sc = wb.create_sheet('Scenarios')
headers = ['Scenario', 'SOM_pct', 'Conv_rate', 'ActiveUsers', 'PayingUsers', 'Pro_pct', 'ARPU', 'Revenue', 'CAC', 'AcquisitionCost', 'GrossProfit', 'NetProfit', 'LTV', 'LTV/CAC']
sc.append(headers)
for c in range(1, len(headers)+1):
    sc.cell(row=1, column=c).font = Font(bold=True)

# Scenario data (name, SOM_pct, Conv_rate, Pro_pct)
scenarios = [
    ('Conservative', 0.02, 0.03, 0.7),
    ('Base', 0.05, 0.05, 0.7),
    ('Optimistic', 0.10, 0.08, 0.7),
]

start_row = 2
for i, (name, som, conv, pro_pct) in enumerate(scenarios, start=start_row):
    sc[f'A{i}'] = name
    sc[f'B{i}'] = som
    sc[f'C{i}'] = conv
    # ActiveUsers = Assumptions!B2 * SOM_pct
    sc[f'D{i}'] = f"=Assumptions!$B$2 * B{i}"
    # PayingUsers = ActiveUsers * Conv_rate
    sc[f'E{i}'] = f"=D{i} * C{i}"
    sc[f'F{i}'] = pro_pct
    # ARPU = price_standard*(1-pro_pct) + price_pro*pro_pct
    sc[f'G{i}'] = f"=Assumptions!$B$3*(1-F{i}) + Assumptions!$B$4*F{i}"
    # Revenue = PayingUsers * ARPU
    sc[f'H{i}'] = f"=E{i} * G{i}"
    # CAC from assumptions
    sc[f'I{i}'] = '=Assumptions!$B$6'
    # AcquisitionCost = PayingUsers * CAC
    sc[f'J{i}'] = f"=E{i} * I{i}"
    # GrossProfit = Revenue * GM
    sc[f'K{i}'] = f"=H{i} * Assumptions!$B$7"
    # NetProfit = GrossProfit - AcquisitionCost - OPS_ANNUAL - DEV_ONE_OFF
    sc[f'L{i}'] = f"=K{i} - J{i} - Assumptions!$B$10 - Assumptions!$B$9"
    # LTV = ARPU * GM * Lifetime
    sc[f'M{i}'] = f"=G{i} * Assumptions!$B$7 * Assumptions!$B$8"
    # LTV/CAC
    sc[f'N{i}'] = f"=M{i} / I{i}"

# Format columns width
for idx, col in enumerate(['A','B','C','D','E','F','G','H','I','J','K','L','M','N'], start=1):
    sc.column_dimensions[col].width = 15

# --- UnitEconomics sheet ---
ue = wb.create_sheet('UnitEconomics')
ue.append(['Metric', 'Formula / Value'])
ue['A1'].font = Font(bold=True)
ue['B1'].font = Font(bold=True)

ue.append(['ARPU example (mix 30% standard,70% pro)', '=0.3*Assumptions!$B$3 + 0.7*Assumptions!$B$4'])
ue.append(['LTV example', '=B2 * Assumptions!$B$7 * Assumptions!$B$8'])
ue.append(['CAC (assumption)', '=Assumptions!$B$6'])
ue.append(['LTV/CAC (example)', '=B3 / B4'])
ue.append(['CAC Payback (months) example', '=Assumptions!$B$6 / ((0.3*Assumptions!$B$3 + 0.7*Assumptions!$B$4)/12)'])

# Adjust column widths
ue.column_dimensions['A'].width = 48
ue.column_dimensions['B'].width = 48

# --- Costs sheet (detailed) ---
costs = wb.create_sheet('Costs')
costs.append(['Cost Category', 'Line Item', 'Annual (CAD)', 'Notes'])
for c in range(1,5):
    costs.cell(row=1, column=c).font = Font(bold=True)

costs_data = [
    ('Development','Frontend devs (salaries/contract)', 180000, 'Year 1 estimate, CAD'),
    ('Development','Backend devs & API', 150000, 'Year 1 estimate, CAD'),
    ('Development','QA / Testing', 30000, 'Year 1 estimate, CAD'),
    ('Development','DevOps / Infra setup', 20000, 'One-time infra setup, CAD'),
    ('Infrastructure','Hosting (annual)', 36000, 'Cloud hosting, scaling'),
    ('Infrastructure','CDN / Maps / APIs', 12000, 'Third-party API costs (maps, data)'),
    ('Marketing','Paid Ads & UA', 90000, 'Annual marketing budget, CAD'),
    ('Marketing','Content & SEO', 18000, 'Ongoing content, CAD'),
    ('Sales','Promotions / Discounts', 12000, 'Promotional credits, CAD'),
    ('Support','Customer support (annual)', 48000, 'Support staff, CAD'),
    ('Legal & Finance','Legal / Compliance', 15000, 'Legal fees, CAD'),
    ('Legal & Finance','Payment processing fees', 12000, 'Estimated fees on revenue'),
    ('Other','Office / Tools / Misc', 10000, 'Accounting, tools, subscriptions'),
    ('One-time','MVP design & research', 25000, 'One-time UX / research costs'),
]

row = 2
for cat, item, amt, note in costs_data:
    costs[f'A{row}'] = cat
    costs[f'B{row}'] = item
    costs[f'C{row}'] = amt
    costs[f'D{row}'] = note
    row += 1

costs.column_dimensions['A'].width = 24
costs.column_dimensions['B'].width = 36
costs.column_dimensions['C'].width = 18
costs.column_dimensions['D'].width = 40

# Add totals row
costs[f'B{row}'] = 'Total Annual Costs'
costs[f'C{row}'] = f"=SUM(C2:C{row-1})"
costs[f'D{row}'] = 'Sum of annual cost lines (excl one-time)'

# Save workbook
out_file = '..\\financial_model_canada.xlsx'
wb.save(out_file)
print(f'Created {out_file}')

# ------------------------------------------------------------------
# Compute numeric scenario outputs and 3-year projections (write as values)
# ------------------------------------------------------------------
assumptions = {k: v for (k, v, _) in ass_data}

# Helper to sum annual cost lines (exclude category 'One-time')
annual_costs = sum(item[2] for item in costs_data if item[0] != 'One-time')

# Add a numeric scenarios sheet
sn = wb.create_sheet('Scenarios_Numeric')
sn_headers = ['Scenario','ActiveUsers','PayingUsers','ARPU','Revenue','AcquisitionCost','GrossProfit','NetProfit','LTV','LTV/CAC']
sn.append(sn_headers)
for c in range(1, len(sn_headers)+1):
    sn.cell(row=1, column=c).font = Font(bold=True)

for name, som, conv, pro_pct in scenarios:
    TAM = assumptions['TAM (total drivers)']
    price_std = assumptions['price_standard (CAD/yr)']
    price_pro = assumptions['price_pro (CAD/yr)']
    CAC = assumptions['CAC (CAD)']
    GM = assumptions['Gross Margin']
    Lifetime = assumptions['Lifetime (years)']
    DEV_ONE_OFF = assumptions['DEV_ONE_OFF (CAD)']
    OPS_ANNUAL = assumptions['OPS_ANNUAL (CAD)']

    active = TAM * som
    paying = active * conv
    arpu = price_std * (1 - pro_pct) + price_pro * pro_pct
    revenue = paying * arpu
    acq_cost = paying * CAC
    gross = revenue * GM
    total_costs = annual_costs + OPS_ANNUAL
    net = gross - acq_cost - total_costs - DEV_ONE_OFF
    ltv = arpu * GM * Lifetime
    ltv_cac = ltv / CAC if CAC else None

    sn.append([name, int(active), int(paying), round(arpu,2), round(revenue,2), round(acq_cost,2), round(gross,2), round(net,2), round(ltv,2), round(ltv_cac,2)])

# 3-year projection sheet
proj = wb.create_sheet('Projection_3yr')
proj.append(['Scenario','Year','PayingUsers','Revenue','TotalCosts','NetProfit'])
for c in range(1,6):
    proj.cell(row=1, column=c).font = Font(bold=True)

years = [1,2,3]
growth_map = {'Conservative': assumptions['growth_conservative'], 'Base': assumptions['growth_base'], 'Optimistic': assumptions['growth_optimistic']}
cost_growth = assumptions['annual_cost_growth']

for name, som, conv, pro_pct in scenarios:
    TAM = assumptions['TAM (total drivers)']
    price_std = assumptions['price_standard (CAD/yr)']
    price_pro = assumptions['price_pro (CAD/yr)']
    CAC = assumptions['CAC (CAD)']
    GM = assumptions['Gross Margin']
    Lifetime = assumptions['Lifetime (years)']
    DEV_ONE_OFF = assumptions['DEV_ONE_OFF (CAD)']
    OPS_ANNUAL = assumptions['OPS_ANNUAL (CAD)']
    growth = growth_map[name]

    paying = int(TAM * som * conv)
    for y in years:
        if y == 1:
            paying_y = paying
        else:
            paying_y = int(round(paying * ((1 + growth) ** (y - 1))))

        arpu = price_std * (1 - pro_pct) + price_pro * pro_pct
        revenue_y = paying_y * arpu
        annual_costs_y = annual_costs * ((1 + cost_growth) ** (y - 1)) + OPS_ANNUAL * ((1 + cost_growth) ** (y - 1))
        if y == 1:
            total_costs_y = annual_costs_y + DEV_ONE_OFF
        else:
            total_costs_y = annual_costs_y

        gross_y = revenue_y * GM
        net_y = gross_y - (paying_y * CAC) - total_costs_y
        proj.append([name, y, int(paying_y), round(revenue_y,2), round(total_costs_y,2), round(net_y,2)])

# Add a simple revenue chart per scenario across 3 years
from openpyxl.chart import LineChart, Reference
chart = LineChart()
chart.title = 'Revenue (3-Year) by Scenario'
chart.y_axis.title = 'Revenue (CAD)'
chart.x_axis.title = 'Year'

# Find projection table coordinates
max_row = proj.max_row
for i, (name, som, conv, pro_pct) in enumerate(scenarios, start=0):
    # Revenue column is column 4
    rev_ref = Reference(proj, min_col=4, min_row=2 + i*3, max_row=2 + i*3 + 2)
    chart.add_data(rev_ref, titles_from_data=False)

# Add category (years) from first scenario rows
cats = Reference(proj, min_col=2, min_row=2, max_row=4)
chart.set_categories(cats)
proj.add_chart(chart, 'H2')

# Save updated workbook
wb.save(out_file)
print(f'Updated and saved {out_file} with numeric sheets and projection')
