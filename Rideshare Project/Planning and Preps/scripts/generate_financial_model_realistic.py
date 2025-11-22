#!/usr/bin/env python3
"""
Generate a 'realistic' Canada financial model using cost-cut recommendations.
Creates `financial_model_canada_realistic.xlsx` with Assumptions, Scenarios_Numeric, Projection_3yr, Costs, and UnitEconomics.
"""
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference

wb = Workbook()

# --- Assumptions sheet (Realistic cuts) ---
ass = wb.active
ass.title = 'Assumptions'
ass.append(['Key', 'Value', 'Notes'])
ass['A1'].font = Font(bold=True)
ass['B1'].font = Font(bold=True)
ass['C1'].font = Font(bold=True)

ass_data = [
    ('Country', 'Canada', 'Target market / country'),
    ('Currency', 'CAD', 'Currency for all values in this sheet'),
    ('TAM (total drivers)', 400000, 'Estimated total addressable market (Canada - gig/active drivers)'),
    ('Default SOM %', 0.05, 'Serviceable obtainable market % (targeted)'),
    ('price_standard (CAD/yr)', 99, 'Standard annual price (CAD) - uplifted'),
    ('price_pro (CAD/yr)', 199, 'Pro annual price (CAD)'),
    ('default_pro_pct', 0.70, 'Default percent of paid users on Pro'),
    ('CAC (CAD)', 30, 'Cost to acquire a paid user (CAD) - reduced via partnerships/PLG'),
    ('Gross Margin', 0.80, 'Gross margin on subscription revenue'),
    ('Lifetime (years)', 2, 'Average paid user lifetime (years)'),
    ('growth_conservative', 0.10, 'Annual growth rate for paying users (conservative)'),
    ('growth_base', 0.30, 'Annual growth rate for paying users (base)'),
    ('growth_optimistic', 0.60, 'Annual growth rate for paying users (optimistic)'),
    ('annual_cost_growth', 0.03, 'Annual growth rate for operating costs'),
    ('DEV_ONE_OFF (CAD)', 100000, 'One-time development cost (Year 1, CAD) - lean MVP'),
    ('OPS_ANNUAL (CAD)', 180000, 'Annual operating cost (CAD) - reduced'),
]

r = 2
for key, val, note in ass_data:
    ass[f'A{r}'] = key
    ass[f'B{r}'] = val
    ass[f'C{r}'] = note
    r += 1

for col in ['A','B','C']:
    ass.column_dimensions[col].width = 32

# --- Costs sheet (detailed, scaled-down) ---
costs = wb.create_sheet('Costs')
costs.append(['Cost Category', 'Line Item', 'Annual (CAD)', 'Notes'])
for c in range(1,5):
    costs.cell(row=1, column=c).font = Font(bold=True)

costs_data = [
    ('Development','Frontend devs (salaries/contract)', 70000, 'Lean hires/contractors'),
    ('Development','Backend devs & API', 40000, 'Small team / contractors'),
    ('Development','QA / Testing', 8000, 'Contracted QA'),
    ('Development','DevOps / Infra setup', 8000, 'Minimal infra setup'),
    ('Infrastructure','Hosting (annual)', 20000, 'Cloud hosting, scaled-down'),
    ('Infrastructure','CDN / Maps / APIs', 8000, 'Select cheaper APIs'),
    ('Marketing','Paid Ads & UA', 30000, 'Reduced paid acquisition'),
    ('Marketing','Content & SEO', 8000, 'Content efforts'),
    ('Sales','Promotions / Discounts', 6000, 'Promotional credits'),
    ('Support','Customer support (annual)', 24000, 'Small support team / automation'),
    ('Legal & Finance','Legal / Compliance', 8000, 'Lean legal budget'),
    ('Legal & Finance','Payment processing fees', 8000, 'Estimated fees'),
    ('Other','Office / Tools / Misc', 5000, 'Remote tools'),
    ('One-time','MVP design & research', 15000, 'Lean UX research'),
]

row = 2
for cat, item, amt, note in costs_data:
    costs[f'A{row}'] = cat
    costs[f'B{row}'] = item
    costs[f'C{row}'] = amt
    costs[f'D{row}'] = note
    row += 1

costs.column_dimensions['A'].width = 24
costs.column_dimensions['B'].width = 36
costs.column_dimensions['C'].width = 18
costs.column_dimensions['D'].width = 40

costs[f'B{row}'] = 'Total Annual Costs'
costs[f'C{row}'] = f"=SUM(C2:C{row-1})"
costs[f'D{row}'] = 'Sum of annual cost lines (excl one-time)'

# --- Scenarios Numeric ---
scenarios = [
    ('Conservative', 0.02, 0.03, 0.7),
    ('Base', 0.05, 0.05, 0.7),
    ('Optimistic', 0.10, 0.08, 0.7),
]

sn = wb.create_sheet('Scenarios_Numeric')
sn_headers = ['Scenario','ActiveUsers','PayingUsers','ARPU','Revenue','AcquisitionCost','GrossProfit','NetProfit','LTV','LTV/CAC']
sn.append(sn_headers)
for c in range(1, len(sn_headers)+1):
    sn.cell(row=1, column=c).font = Font(bold=True)

# extract values from assumptions dict-like
ass_map = {k: v for (k, v, _) in ass_data}

annual_costs = sum(item[2] for item in costs_data if item[0] != 'One-time')

for name, som, conv, pro_pct in scenarios:
    TAM = ass_map['TAM (total drivers)']
    price_std = ass_map['price_standard (CAD/yr)']
    price_pro = ass_map['price_pro (CAD/yr)']
    CAC = ass_map['CAC (CAD)']
    GM = ass_map['Gross Margin']
    Lifetime = ass_map['Lifetime (years)']
    DEV_ONE_OFF = ass_map['DEV_ONE_OFF (CAD)']
    OPS_ANNUAL = ass_map['OPS_ANNUAL (CAD)']

    active = int(TAM * som)
    paying = int(active * conv)
    arpu = price_std * (1 - pro_pct) + price_pro * pro_pct
    revenue = paying * arpu
    acq_cost = paying * CAC
    gross = revenue * GM
    total_costs = annual_costs + OPS_ANNUAL
    net = gross - acq_cost - total_costs - DEV_ONE_OFF
    ltv = arpu * GM * Lifetime
    ltv_cac = ltv / CAC if CAC else None

    sn.append([name, active, paying, round(arpu,2), round(revenue,2), round(acq_cost,2), round(gross,2), round(net,2), round(ltv,2), round(ltv_cac,2)])

# --- 3-year projection ---
proj = wb.create_sheet('Projection_3yr')
proj.append(['Scenario','Year','PayingUsers','Revenue','TotalCosts','NetProfit'])
for c in range(1,6):
    proj.cell(row=1, column=c).font = Font(bold=True)

years = [1,2,3]
growth_map = {'Conservative': ass_map['growth_conservative'], 'Base': ass_map['growth_base'], 'Optimistic': ass_map['growth_optimistic']}
cost_growth = ass_map['annual_cost_growth']

for name, som, conv, pro_pct in scenarios:
    TAM = ass_map['TAM (total drivers)']
    price_std = ass_map['price_standard (CAD/yr)']
    price_pro = ass_map['price_pro (CAD/yr)']
    CAC = ass_map['CAC (CAD)']
    GM = ass_map['Gross Margin']
    Lifetime = ass_map['Lifetime (years)']
    DEV_ONE_OFF = ass_map['DEV_ONE_OFF (CAD)']
    OPS_ANNUAL = ass_map['OPS_ANNUAL (CAD)']
    growth = growth_map[name]

    paying = int(TAM * som * conv)
    for y in years:
        if y == 1:
            paying_y = paying
        else:
            paying_y = int(round(paying * ((1 + growth) ** (y - 1))))

        arpu = price_std * (1 - pro_pct) + price_pro * pro_pct
        revenue_y = paying_y * arpu
        annual_costs_y = annual_costs * ((1 + cost_growth) ** (y - 1)) + OPS_ANNUAL * ((1 + cost_growth) ** (y - 1))
        if y == 1:
            total_costs_y = annual_costs_y + DEV_ONE_OFF
        else:
            total_costs_y = annual_costs_y

        gross_y = revenue_y * GM
        net_y = gross_y - (paying_y * CAC) - total_costs_y
        proj.append([name, y, int(paying_y), round(revenue_y,2), round(total_costs_y,2), round(net_y,2)])

# Add a revenue chart
chart = LineChart()
chart.title = 'Revenue (3-Year) by Scenario - Realistic'
chart.y_axis.title = 'Revenue (CAD)'
chart.x_axis.title = 'Year'

for i, (name, som, conv, pro_pct) in enumerate(scenarios, start=0):
    rev_ref = Reference(proj, min_col=4, min_row=2 + i*3, max_row=2 + i*3 + 2)
    chart.add_data(rev_ref, titles_from_data=False)

cats = Reference(proj, min_col=2, min_row=2, max_row=4)
chart.set_categories(cats)
proj.add_chart(chart, 'H2')

out_file = '..\\financial_model_canada_realistic.xlsx'
wb.save(out_file)
print('Created', out_file)
