"""
Read ..\break_even_results.xlsx TopResults and create a Top20 sheet with a bar chart showing Net_Y3_after_DEV.
Writes: ..\break_even_with_chart.xlsx
"""
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

src = '..\\break_even_results.xlsx'
out = '..\\break_even_with_chart.xlsx'

wb_src = load_workbook(src, data_only=True)
if 'TopResults' not in wb_src.sheetnames:
    raise SystemExit('TopResults sheet not found in break_even_results.xlsx')
ws = wb_src['TopResults']

# Read header indices
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data_rows = rows[1:]

# Take top 20 (already sorted by net desc in the generator)
top = data_rows[:20]

# Build labels and nets
labels = []
nets = []
for r in top:
    net = r[0]
    DEV = r[1]
    CAC = r[2]
    conv = r[3]
    pstd = r[4]
    ppro = r[5]
    paying = r[6]
    rev = r[7]
    label = f'DEV{DEV}|CAC{CAC}|c{conv:.0%}|p{pstd}'
    labels.append(label)
    nets.append(net)

# Create output workbook with Top20 sheet
wb = Workbook()
ws_out = wb.active
ws_out.title = 'Top20'
ws_out.append(['Label','Net_Y3_after_DEV'])
ws_out['A1'].font = Font(bold=True)
ws_out['B1'].font = Font(bold=True)

for lab, net in zip(labels, nets):
    ws_out.append([lab, net])

# Create a bar chart
chart = BarChart()
chart.type = 'col'
chart.title = 'Net (Year 3 after DEV) - Top 20 Sweep Results'
chart.y_axis.title = 'Net (CAD)'
chart.x_axis.title = 'Combination (DEV|CAC|conv|price)'

data = Reference(ws_out, min_col=2, min_row=1, max_row=1 + len(nets))
cats = Reference(ws_out, min_col=1, min_row=2, max_row=1 + len(nets))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4

ws_out.add_chart(chart, 'D2')

wb.save(out)
print('Created', out)
