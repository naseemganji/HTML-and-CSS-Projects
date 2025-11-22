#!/usr/bin/env python3
"""
Generate a Canada financial model that includes enterprise/B2B revenue lines.
Output: ..\financial_model_canada_enterprise.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference

wb = Workbook()

# Assumptions (based on realistic scenario but with enterprise revenue)
ass = wb.active
ass.title = 'Assumptions'
ass.append(['Key','Value','Notes'])
for c in range(1,4):
    ass.cell(row=1, column=c).font = Font(bold=True)

ass_data = [
    ('Country','Canada',''),
    ('Currency','CAD',''),
    ('TAM (total drivers)',400000,'Drivers'),
    ('SOM',0.05,''),
    ('price_standard',99,''),
    ('price_pro',199,''),
    ('pro_pct',0.70,''),
    ('CAC',30,''),
    ('GM',0.80,''),
    ('Lifetime',2,''),
    ('DEV_ONE_OFF',100000,''),
    ('OPS_ANNUAL',180000,''),
    ('growth_base',0.30,''),
]

r=2
for k,v,n in ass_data:
    ass[f'A{r}']=k
    ass[f'B{r}']=v
    ass[f'C{r}']=n
    r+=1

for col in ['A','B','C']:
    ass.column_dimensions[col].width = 28

# Costs (simplified - reuse realistic)
costs = wb.create_sheet('Costs')
costs.append(['Category','Line','Annual','Notes'])
for c in range(1,5):
    costs.cell(row=1, column=c).font = Font(bold=True)

costs_data = [
    ('Dev','Team & contractors',110000,''),
    ('Infra','Hosting & APIs',28000,''),
    ('Marketing','Paid & content',38000,''),
    ('Support','Support',24000,''),
    ('Legal','Legal & payments',16000,''),
    ('Other','Tools & misc',5000,''),
]
row=2
for cat, line, amt, note in costs_data:
    costs[f'A{row}']=cat
    costs[f'B{row}']=line
    costs[f'C{row}']=amt
    costs[f'D{row}']=note
    row+=1

costs[f'B{row}']='Total Annual Costs'
costs[f'C{row}']=f"=SUM(C2:C{row-1})"

# Enterprise revenue assumptions
# We'll model a small enterprise pilot scaling: Year1: 1 contract @ 12k, Year2: 3 @12k, Year3: 8 @12k
enterprise = wb.create_sheet('Enterprise')
enterprise.append(['Year','Contracts','Revenue_per_contract','Total_Enterprise_Revenue'])
for c in range(1,5):
    enterprise.cell(row=1, column=c).font = Font(bold=True)

enterprise_data = [(1,1,12000,12000),(2,3,12000,36000),(3,8,12000,96000)]
for y, cnt, rpc, total in enterprise_data:
    enterprise.append([y,cnt,rpc,total])

# Scenarios numeric (subscription revenue) - reusing realistic convs
scenarios = [('Conservative',0.02,0.03,0.7),('Base',0.05,0.05,0.7),('Optimistic',0.10,0.08,0.7)]

sn = wb.create_sheet('Scenarios_Numeric')
headers=['Scenario','ActiveUsers','PayingUsers','ARPU','Revenue_subs','Enterprise_Revenue_Year1','Revenue_total','AcqCost','GrossProfit','NetProfit']
for c,h in enumerate(headers, start=1):
    sn.cell(row=1, column=c).value = h
    sn.cell(row=1, column=c).font = Font(bold=True)

# compute subscription revenue per scenario (year1), and include enterprise Y1 revenue
ass_map = {k:v for (k,v,_) in ass_data}
annual_costs = sum(item[2] for item in costs_data)
enterprise_y1 = enterprise_data[0][3]

for name, som, conv, pro_pct in scenarios:
    TAM = ass_map['TAM (total drivers)']
    active = int(TAM * som)
    paying = int(active * conv)
    arpu = ass_map['price_standard']*(1-pro_pct) + ass_map['price_pro']*pro_pct
    revenue_subs = paying * arpu
    revenue_total = revenue_subs + enterprise_y1
    acq = paying * ass_map['CAC']
    gross = revenue_total * ass_map['GM']
    net = gross - acq - annual_costs - ass_map['DEV_ONE_OFF'] - ass_map['OPS_ANNUAL']
    sn.append([name, active, paying, round(arpu,2), round(revenue_subs,2), enterprise_y1, round(revenue_total,2), round(acq,2), round(gross,2), round(net,2)])

# Projection 3yr including enterprise revenue per year from sheet
proj = wb.create_sheet('Projection_3yr')
proj.append(['Scenario','Year','PayingUsers','Revenue_subs','Enterprise_Revenue','Revenue_total','TotalCosts','NetProfit'])
for c in range(1,9):
    proj.cell(row=1, column=c).font = Font(bold=True)

years=[1,2,3]
growth_map={'Conservative':ass_map['growth_base']*0.33,'Base':ass_map['growth_base'],'Optimistic':ass_map['growth_base']*2}
enterprise_by_year={1:enterprise_data[0][3],2:enterprise_data[1][3],3:enterprise_data[2][3]}

for name, som, conv, pro_pct in scenarios:
    TAM = ass_map['TAM (total drivers)']
    paying0 = int(TAM * som * conv)
    growth = growth_map[name]
    for y in years:
        paying_y = int(round(paying0 * ((1+growth)**(y-1))))
        arpu = ass_map['price_standard']*(1-pro_pct) + ass_map['price_pro']*pro_pct
        rev_subs = paying_y * arpu
        ent_rev = enterprise_by_year[y]
        rev_total = rev_subs + ent_rev
        annual_costs_y = annual_costs * ((1+0.03)**(y-1)) + ass_map['OPS_ANNUAL'] * ((1+0.03)**(y-1))
        if y==1:
            total_costs = annual_costs_y + ass_map['DEV_ONE_OFF']
        else:
            total_costs = annual_costs_y
        gross = rev_total * ass_map['GM']
        net = gross - (paying_y * ass_map['CAC']) - total_costs
        proj.append([name, y, paying_y, round(rev_subs,2), ent_rev, round(rev_total,2), round(total_costs,2), round(net,2)])

# revenue chart
chart = LineChart()
chart.title='Revenue (3-Year) including Enterprise'
chart.y_axis.title='Revenue (CAD)'
chart.x_axis.title='Year'
for i, (name, som, conv, _) in enumerate(scenarios, start=0):
    rev_ref = Reference(proj, min_col=6, min_row=2 + i*3, max_row=2 + i*3 + 2)
    chart.add_data(rev_ref, titles_from_data=False)
cats = Reference(proj, min_col=2, min_row=2, max_row=4)
chart.set_categories(cats)
proj.add_chart(chart, 'I2')

out='..\\financial_model_canada_enterprise.xlsx'
wb.save(out)
print('Created', out)
