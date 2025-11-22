from openpyxl import load_workbook
wb=load_workbook('..\\financial_model_canada_sensitivities.xlsx', data_only=True)
print('\n--- Year3 Comparison ---')
ws=wb['Year3_Comparison']
for r in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=5, values_only=True):
    print(r)
print('\n--- Summary (sample rows) ---')
ws2=wb['Summary']
for r in ws2.iter_rows(min_row=1, max_row=20, min_col=1, max_col=7, values_only=True):
    print(r)
