"""
Create Top20 chart for constrained break-even results and save as ..\break_even_constrained_chart.xlsx
"""
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

src='..\\break_even_constrained.xlsx'
out='..\\break_even_constrained_chart.xlsx'
wb_src = load_workbook(src, data_only=True)
ws = wb_src['TopResults']
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data = rows[1:21]

wb = Workbook()
ws_out = wb.active
ws_out.title='Top20'
ws_out.append(['Label','Net_Y3_after_DEV'])
ws_out['A1'].font = Font(bold=True)
ws_out['B1'].font = Font(bold=True)

for r in data:
    net, DEV, CAC, conv, pstd, ppro, paying, rev = r
    label = f'DEV{DEV}|CAC{CAC}|c{conv}|p{pstd}'
    ws_out.append([label, net])

chart = BarChart()
chart.title='Constrained Sweep - Top 20 Nets (Year3 after DEV)'
chart.y_axis.title='Net (CAD)'
chart.x_axis.title='Combo'

data_ref = Reference(ws_out, min_col=2, min_row=1, max_row=1+len(data))
cat_ref = Reference(ws_out, min_col=1, min_row=2, max_row=1+len(data))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)
ws_out.add_chart(chart, 'D2')

wb.save(out)
print('Created', out)
