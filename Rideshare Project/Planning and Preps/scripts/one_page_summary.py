"""
Create a one-page PNG and PDF comparing Year3 Net for Baseline (original), Realistic, Enterprise, and Constrained best.
Requires: matplotlib, reportlab
"""
import math
from openpyxl import load_workbook

# Gather numbers
wb_real = load_workbook('..\\financial_model_canada_realistic.xlsx', data_only=True)
pr = wb_real['Projection_3yr']
real = None
for r in pr.iter_rows(values_only=True):
    if r[0]=='Base' and r[1]==3:
        real = float(r[5])

wb_ent = load_workbook('..\\financial_model_canada_enterprise.xlsx', data_only=True)
pe = wb_ent['Projection_3yr']
ent = None
for r in pe.iter_rows(values_only=True):
    if r[0]=='Base' and r[1]==3:
        ent = float(r[7])

wb_con = load_workbook('..\\break_even_constrained.xlsx', data_only=True)
tr = wb_con['TopResults']
con_best = float(tr['A2'].value)

labels = ['Realistic','Enterprise','ConstrainedBest']
values = [real, ent, con_best]

# Plot using matplotlib
import matplotlib.pyplot as plt
import numpy as np

plt.style.use('classic')
fig, ax = plt.subplots(figsize=(8,5))
colors = ['#d9534f' if v<0 else '#5cb85c' for v in values]
ax.bar(labels, values, color=colors)
ax.set_title('Year 3 Net Profit Comparison')
ax.axhline(0, color='black', linewidth=0.8)
for i,v in enumerate(values):
    ax.text(i, v + (20000 if v>=0 else -50000), f'{v:,.0f} CAD', ha='center', va='bottom' if v>=0 else 'top')

plt.tight_layout()
png_out = '..\\one_page_summary.png'
plt.savefig(png_out)
print('Wrote', png_out)

# Create a simple PDF embedding the PNG via reportlab
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

pdf_out = '..\\one_page_summary.pdf'
c = canvas.Canvas(pdf_out, pagesize=letter)
img = ImageReader(png_out)
c.drawImage(img, 72, 300, width=468, height=300)
# add a short text summary
c.setFont('Helvetica-Bold', 12)
c.drawString(72, 280, 'DriveGo - Year 3 Net Profit Comparison (selected scenarios)')
c.setFont('Helvetica', 10)
c.drawString(72, 260, f'Realistic Year3 Net: {real:,.0f} CAD')
c.drawString(72, 245, f'Enterprise Year3 Net: {ent:,.0f} CAD')
c.drawString(72, 230, f'Constrained Best Net: {con_best:,.0f} CAD')
c.save()
print('Wrote', pdf_out)
