"""Run sensitivity scenarios for the Canada financial model and write a summary workbook.
Produces: ../financial_model_canada_sensitivities.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font

# Shared costs data (same as generator)
costs_data = [
    ('Development','Frontend devs (salaries/contract)', 180000),
    ('Development','Backend devs & API', 150000),
    ('Development','QA / Testing', 30000),
    ('Development','DevOps / Infra setup', 20000),
    ('Infrastructure','Hosting (annual)', 36000),
    ('Infrastructure','CDN / Maps / APIs', 12000),
    ('Marketing','Paid Ads & UA', 90000),
    ('Marketing','Content & SEO', 18000),
    ('Sales','Promotions / Discounts', 12000),
    ('Support','Customer support (annual)', 48000),
    ('Legal & Finance','Legal / Compliance', 15000),
    ('Legal & Finance','Payment processing fees', 12000),
    ('Other','Office / Tools / Misc', 10000),
    ('One-time','MVP design & research', 25000),
]

annual_costs_base = sum(item[2] for item in costs_data if item[0] != 'One-time')

# Baseline assumptions (matching generator)
baseline = {
    'Country': 'Canada',
    'Currency': 'CAD',
    'TAM': 400000,
    'SOM': 0.05,
    'price_standard': 79,
    'price_pro': 149,
    'pro_pct': 0.70,
    'CAC': 70,
    'GM': 0.80,
    'Lifetime': 2,
    'DEV_ONE_OFF': 450000,
    'OPS_ANNUAL': 300000,
    'growth': {'Conservative':0.10,'Base':0.30,'Optimistic':0.60},
    'annual_cost_growth': 0.03,
}

scenarios_def = {
    'Baseline': baseline,
    'MVP_Lean': dict(baseline, **{'DEV_ONE_OFF':100000, 'OPS_ANNUAL':180000}),
    'LowCAC_HigherConv': dict(baseline, **{'CAC':30, 'conv_base':0.10}),
    'Price_Uplift': dict(baseline, **{'price_standard':99, 'price_pro':199}),
}

# Combined scenario: MVP lean + low CAC + higher conv + price uplift
combined = dict(baseline)
combined.update({'DEV_ONE_OFF':100000,'OPS_ANNUAL':180000,'CAC':30,'price_standard':99,'price_pro':199,'conv_base':0.10})
scenarios_def['Combined'] = combined

# Default conv rates per scenario (if not provided, use generator defaults)
def get_conv(name, base_conv):
    # generator used convs: Conservative 0.03, Base 0.05, Optimistic 0.08
    return base_conv

# We'll use three internal scenario types: Conservative, Base, Optimistic conv rates
conv_rates = {'Conservative':0.03,'Base':0.05,'Optimistic':0.08}

wb = Workbook()
summary = wb.active
summary.title = 'Summary'
summary.append(['Scenario','ConvType','Year','PayingUsers','Revenue','TotalCosts','NetProfit'])
for c in range(1,7):
    summary.cell(row=1, column=c).font = Font(bold=True)

# For each defined sensitivity, compute 3-year projections for each conv type
for sname, asum in scenarios_def.items():
    TAM = asum['TAM']
    SOM = asum['SOM']
    price_std = asum['price_standard']
    price_pro = asum['price_pro']
    pro_pct = asum['pro_pct']
    CAC = asum['CAC']
    GM = asum['GM']
    Lifetime = asum['Lifetime']
    DEV_ONE_OFF = asum['DEV_ONE_OFF']
    OPS_ANNUAL = asum['OPS_ANNUAL']
    annual_costs = annual_costs_base
    cost_growth = asum.get('annual_cost_growth', 0.03)
    base_conv_override = asum.get('conv_base', None)
    growth_map = asum.get('growth', {'Conservative':0.10,'Base':0.30,'Optimistic':0.60})

    for conv_type, conv_rate in conv_rates.items():
        if base_conv_override is not None and conv_type == 'Base':
            conv = base_conv_override
        else:
            conv = conv_rate

        paying_start = int(TAM * SOM * conv)
        for year in [1,2,3]:
            if year == 1:
                paying_y = paying_start
            else:
                growth = growth_map.get(conv_type, 0.30)
                paying_y = int(round(paying_start * ((1 + growth) ** (year - 1))))

            arpu = price_std * (1 - pro_pct) + price_pro * pro_pct
            revenue_y = paying_y * arpu
            annual_costs_y = annual_costs * ((1 + cost_growth) ** (year - 1)) + OPS_ANNUAL * ((1 + cost_growth) ** (year - 1))
            if year == 1:
                total_costs_y = annual_costs_y + DEV_ONE_OFF
            else:
                total_costs_y = annual_costs_y
            gross_y = revenue_y * GM
            net_y = gross_y - (paying_y * CAC) - total_costs_y

            summary.append([sname, conv_type, year, paying_y, round(revenue_y,2), round(total_costs_y,2), round(net_y,2)])

# Also write a per-scenario sheet with Year3 net profit for quick comparison
comp = wb.create_sheet('Year3_Comparison')
comp.append(['Scenario','ConvType','Year3_PayingUsers','Year3_Revenue','Year3_NetProfit'])
comp.cell(row=1, column=1).font = Font(bold=True)
comp.cell(row=1, column=2).font = Font(bold=True)
comp.cell(row=1, column=3).font = Font(bold=True)
comp.cell(row=1, column=4).font = Font(bold=True)
comp.cell(row=1, column=5).font = Font(bold=True)

# Read summary rows to fill comparison
for row in summary.iter_rows(min_row=2, values_only=True):
    sname, conv_type, year, paying, revenue, costs, net = row
    if year == 3:
        comp.append([sname, conv_type, paying, revenue, net])

out = '..\\financial_model_canada_sensitivities.xlsx'
wb.save(out)
print('Created', out)
