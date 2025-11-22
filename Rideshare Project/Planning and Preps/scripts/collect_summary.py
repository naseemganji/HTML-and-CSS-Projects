from openpyxl import load_workbook

def get_realistic_year3():
    wb=load_workbook('..\\financial_model_canada_realistic.xlsx', data_only=True)
    pr=wb['Projection_3yr']
    for r in pr.iter_rows(values_only=True):
        if r[0]=='Base' and r[1]==3:
            return r[5]
    return None

def get_enterprise_year3():
    wb=load_workbook('..\\financial_model_canada_enterprise.xlsx', data_only=True)
    pr=wb['Projection_3yr']
    for r in pr.iter_rows(values_only=True):
        if r[0]=='Base' and r[1]==3:
            return r[7]
    return None

def get_constrained_best():
    wb=load_workbook('..\\break_even_constrained.xlsx', data_only=True)
    tr=wb['TopResults']
    # second row first col
    return tr['A2'].value

print('Realistic Year3 Net:', get_realistic_year3())
print('Enterprise Year3 Net:', get_enterprise_year3())
print('Constrained best Net:', get_constrained_best())
