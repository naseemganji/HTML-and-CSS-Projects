from openpyxl import load_workbook
wb=load_workbook('..\\financial_model_canada.xlsx', data_only=True)
print('\n--- Scenarios_Numeric (rows) ---')
ws=wb['Scenarios_Numeric']
for r in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
    print(r)
print('\n--- Projection_3yr (top rows) ---')
wp=wb['Projection_3yr']
for r in wp.iter_rows(min_row=1, max_row=12, min_col=1, max_col=6, values_only=True):
    print(r)
